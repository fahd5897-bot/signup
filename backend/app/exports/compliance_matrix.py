"""The compliance matrix — the artefact a procurement authority audits.

Every requirement in the tender, the answer given, who signed it off, and which
of the bidder's documents back it. This is the file that has to survive a
challenge months after submission, so it reports the *true* state of every row
including the ones that were never answered. A matrix that quietly omits
unanswered optional items reads as complete and is worse than useless.
"""

from __future__ import annotations

import io
from typing import Final

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.db.models.enums import ProposalStatus
from app.exports.models import ExportBundle, ExportRow

MEDIA_TYPE: Final = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_HEADERS_EN: Final = (
    "Reference",
    "Requirement",
    "Mandatory",
    "Response",
    "Status",
    "Source documents",
    "Citations",
    "Reviewed by",
    "Reviewed at",
    "Human edited",
)
_HEADERS_AR: Final = (
    "المرجع",
    "المتطلب",
    "إلزامي",
    "الرد",
    "الحالة",
    "المستندات المصدر",
    "الاستشهادات",
    "المراجع",
    "تاريخ المراجعة",
    "تحرير بشري",
)

#: Column widths in characters, index-aligned with the headers above.
_WIDTHS: Final = (16, 60, 11, 80, 16, 34, 11, 26, 20, 13)

_HEADER_FILL: Final = PatternFill("solid", fgColor="1F2937")
#: Amber, for a requirement carrying no citation. Not an error — a reviewer may
#: legitimately have written the answer themselves — but an auditor should be
#: able to find every one of them at a glance.
_UNCITED_FILL: Final = PatternFill("solid", fgColor="FEF3C7")
#: Red, for a mandatory requirement with no answer at all.
_MISSING_FILL: Final = PatternFill("solid", fgColor="FEE2E2")


def build_compliance_matrix(bundle: ExportBundle) -> bytes:
    """Render the matrix as XLSX bytes.

    XLSX rather than CSV because the matrix is read by people: it needs column
    widths, wrapped text, and — for Arabic tenders — a right-to-left sheet, none
    of which survive a CSV.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Compliance Matrix"

    # Reverses column order and the freeze pane, so an Arabic reader starts at
    # the reference column instead of ending on it.
    sheet.sheet_view.rightToLeft = bundle.is_rtl

    headers = _HEADERS_AR if bundle.is_rtl else _HEADERS_EN
    _write_header(sheet, headers)

    for row in bundle.rows:
        sheet.append(_to_cells(row, rtl=bundle.is_rtl))
        _shade(sheet, row)

    _finish(sheet, len(headers), bundle.is_rtl)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# ------------------------------------------------------------------ internals
def _write_header(sheet: Worksheet, headers: tuple[str, ...]) -> None:
    sheet.append(list(headers))
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)


def _to_cells(row: ExportRow, *, rtl: bool) -> list[object]:
    yes, no = ("نعم", "لا") if rtl else ("Yes", "No")
    return [
        row.requirement_ref,
        row.requirement_text,
        yes if row.is_mandatory else no,
        row.answer_text or "",
        row.status.value,
        "\n".join(row.source_documents),
        row.citation_count,
        row.reviewer or "",
        # Naive local time, because Excel has no timezone-aware datetime and
        # writing an aware one raises. UTC is what the database stores.
        row.reviewed_at.replace(tzinfo=None) if row.reviewed_at else "",
        yes if row.was_edited_by_human else no,
    ]


def _shade(sheet: Worksheet, row: ExportRow) -> None:
    """Colour the rows an auditor needs to find first."""
    fill = None
    if row.is_mandatory and not row.is_answered:
        fill = _MISSING_FILL
    elif row.is_answered and row.citation_count == 0:
        fill = _UNCITED_FILL
    if fill is None:
        return
    for cell in sheet[sheet.max_row]:
        cell.fill = fill


def _finish(sheet: Worksheet, column_count: int, rtl: bool) -> None:
    for index in range(1, column_count + 1):
        sheet.column_dimensions[get_column_letter(index)].width = _WIDTHS[index - 1]

    alignment = Alignment(
        wrap_text=True,
        vertical="top",
        # Excel's own "right-to-left" reading order for the cell, so mixed
        # Arabic and Latin text (a standard reference like ISO 27001 inside an
        # Arabic sentence) lays out the way a reader expects.
        readingOrder=2 if rtl else 1,
        horizontal="right" if rtl else "left",
    )
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = alignment

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(column_count)}{sheet.max_row}"


def summarise(bundle: ExportBundle) -> dict[str, int]:
    """Counts for the export receipt shown to the bid manager."""
    return {
        "requirements": len(bundle.rows),
        "answered": sum(1 for r in bundle.rows if r.is_answered),
        "mandatory": sum(1 for r in bundle.rows if r.is_mandatory),
        "uncited": sum(1 for r in bundle.rows if r.is_answered and r.citation_count == 0),
        "exported": sum(1 for r in bundle.rows if r.status is ProposalStatus.EXPORTED),
    }


__all__ = ["MEDIA_TYPE", "build_compliance_matrix", "summarise"]
