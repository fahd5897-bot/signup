"""Export, against live PostgreSQL with RLS enforced.

The last gate, and the only one that produces a file. Every test here asserts
something that must not be possible: exporting with a mandatory requirement
unapproved, an unapproved draft appearing in the submitted document, or an
export leaving the answers it shipped indistinguishable from ones it did not.
"""

from __future__ import annotations

import io
import uuid
import zipfile

import pytest
from app.db.models.enums import GroundingVerdict, Language, ProposalStatus, UserRole
from app.db.repositories.proposals import ProposalRepository
from app.db.session import tenant_session
from app.services.export_service import ExportBlockedError, ExportFormat, ExportService
from app.services.review_service import ReviewService
from openpyxl import load_workbook
from sqlalchemy import text

pytestmark = pytest.mark.integration

CITED = [
    {"chunk_id": "c1", "document_name": "iso-27001.pdf", "page_number": 4},
    {"chunk_id": "c2", "document_name": "iso-27001.pdf", "page_number": 5},
    {"chunk_id": "c3", "document_name": "company-profile.docx"},
]


@pytest.fixture(autouse=True)
def _app_role(monkeypatch, app_dsn):
    monkeypatch.setenv("POSTGRES_DSN", app_dsn)
    import app.db.session as session_module
    from app.core.config import get_settings

    def _reset() -> None:
        get_settings.cache_clear()
        session_module._engine = None
        session_module._session_factory = None

    _reset()
    yield
    _reset()


@pytest.fixture
async def workspace(superuser_engine, two_tenants):
    a, _ = two_tenants
    workspace_id = uuid.uuid4()
    async with superuser_engine.begin() as conn:
        await conn.execute(text("SELECT set_config('app.tenant_id', :t, true)"), {"t": str(a)})
        owner_id = (
            await conn.execute(text("SELECT id FROM users WHERE tenant_id = :t LIMIT 1"), {"t": a})
        ).scalar_one()
        await conn.execute(
            text(
                "INSERT INTO workspaces (id,tenant_id,name,status,response_language,"
                "owner_id,grounding_config,requirements_total,requirements_approved) "
                "VALUES (:w,:t,'Ministry of Health 2026','draft','en',:o,'{}',0,0)"
            ),
            {"w": workspace_id, "t": a, "o": owner_id},
        )
    return a, workspace_id, owner_id


async def _proposal(
    tenant_id: uuid.UUID,
    workspace_id: uuid.UUID,
    *,
    ref: str,
    answer: str | None = "We hold ISO 27001, certificate 12345.",
    citations: list[dict] | None = None,
    is_mandatory: bool = True,
) -> uuid.UUID:
    async with tenant_session(tenant_id) as session:
        proposal = await ProposalRepository(session).save_generation(
            tenant_id=tenant_id,
            workspace_id=workspace_id,
            requirement_ref=ref,
            requirement_text=f"Requirement {ref}",
            answer_text=answer,
            language=Language.EN,
            status=ProposalStatus.DRAFT,
            citations=CITED if citations is None else citations,
            retrieved_chunk_ids=["c1"],
            grounding_verdict=GroundingVerdict.VERIFIED,
            citation_coverage=1.0,
            top_retrieval_score=0.93,
            confidence_score=0.94,
            abstention_reason=None,
            model_id="claude-opus-5",
            prompt_version="answer-gen/2026-08-17",
            is_mandatory=is_mandatory,
        )
        return proposal.id


async def _approve(tenant_id, proposal_id, owner_id, version: int = 1) -> None:
    await ReviewService().apply_action(
        tenant_id=tenant_id,
        proposal_id=proposal_id,
        actor_id=owner_id,
        actor_role=UserRole.OWNER,
        action=ProposalStatus.APPROVED,
        expected_version=version,
        review_notes="checked",
    )


# ------------------------------------------------------------------ the gate
async def test_export_refuses_while_a_mandatory_requirement_is_unapproved(workspace):
    tenant_id, workspace_id, owner_id = workspace
    first = await _proposal(tenant_id, workspace_id, ref="1.1")
    await _proposal(tenant_id, workspace_id, ref="1.2")
    await _approve(tenant_id, first, owner_id)

    with pytest.raises(ExportBlockedError) as excinfo:
        await ExportService().export(
            tenant_id=tenant_id,
            workspace_id=workspace_id,
            export_format=ExportFormat.DOCX,
        )
    assert excinfo.value.context["outstanding"] == 1
    assert "1.2" in excinfo.value.context["examples"]


async def test_an_empty_workspace_cannot_be_exported(workspace):
    """Nothing answered means nothing signed off; a zero-requirement export
    would be an empty submission that looks successful."""
    tenant_id, workspace_id, _ = workspace
    with pytest.raises(ExportBlockedError):
        await ExportService().export(
            tenant_id=tenant_id,
            workspace_id=workspace_id,
            export_format=ExportFormat.MATRIX,
        )


async def test_optional_requirements_do_not_block_the_gate(workspace):
    tenant_id, workspace_id, owner_id = workspace
    mandatory = await _proposal(tenant_id, workspace_id, ref="1.1")
    await _proposal(tenant_id, workspace_id, ref="1.9", is_mandatory=False)
    await _approve(tenant_id, mandatory, owner_id)

    artifact = await ExportService().export(
        tenant_id=tenant_id, workspace_id=workspace_id, export_format=ExportFormat.MATRIX
    )
    assert artifact.summary["requirements"] == 2
    assert artifact.summary["answered"] == 1


# ------------------------------------------------------------- what ships
async def test_an_unapproved_answer_never_reaches_the_document(workspace):
    """The whole point of the gate. The optional draft is real, plausible text
    that nobody signed off, and it must not be in the file."""
    tenant_id, workspace_id, owner_id = workspace
    approved = await _proposal(tenant_id, workspace_id, ref="1.1", answer="Approved answer.")
    await _proposal(
        tenant_id,
        workspace_id,
        ref="1.9",
        answer="Unreviewed draft that must not ship.",
        is_mandatory=False,
    )
    await _approve(tenant_id, approved, owner_id)

    artifact = await ExportService().export(
        tenant_id=tenant_id, workspace_id=workspace_id, export_format=ExportFormat.DOCX
    )
    body = _docx_text(artifact.content)
    assert "Approved answer." in body
    assert "must not ship" not in body


async def test_the_matrix_reports_unanswered_rows_rather_than_hiding_them(workspace):
    """A matrix that omits what was not answered reads as complete, which is
    the one thing an audit artefact must never do."""
    tenant_id, workspace_id, owner_id = workspace
    approved = await _proposal(tenant_id, workspace_id, ref="1.1")
    await _proposal(tenant_id, workspace_id, ref="1.9", is_mandatory=False)
    await _approve(tenant_id, approved, owner_id)

    artifact = await ExportService().export(
        tenant_id=tenant_id, workspace_id=workspace_id, export_format=ExportFormat.MATRIX
    )
    sheet = load_workbook(io.BytesIO(artifact.content)).active
    refs = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True)]
    assert refs == ["1.1", "1.9"]

    by_ref = {row[0]: row for row in sheet.iter_rows(min_row=2, values_only=True)}
    # openpyxl reads an empty cell back as None, so check for absence of text
    # rather than for a particular empty representation.
    assert not by_ref["1.9"][3], "the unapproved draft's text must not appear"
    assert by_ref["1.1"][3]


async def test_the_matrix_names_the_reviewer_and_the_source_documents(workspace):
    """The columns a procurement authority actually audits."""
    tenant_id, workspace_id, owner_id = workspace
    proposal_id = await _proposal(tenant_id, workspace_id, ref="1.1")
    await _approve(tenant_id, proposal_id, owner_id)

    artifact = await ExportService().export(
        tenant_id=tenant_id, workspace_id=workspace_id, export_format=ExportFormat.MATRIX
    )
    sheet = load_workbook(io.BytesIO(artifact.content)).active
    row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))

    assert row[5].splitlines() == ["iso-27001.pdf", "company-profile.docx"], "deduped, in order"
    assert row[6] == 3, "citation count"
    assert row[7] == "acme", "the named human who approved it"
    assert row[8] is not None, "and when"


# -------------------------------------------------------------- after export
async def test_exported_answers_are_marked_and_can_no_longer_be_re_reviewed(workspace):
    tenant_id, workspace_id, owner_id = workspace
    proposal_id = await _proposal(tenant_id, workspace_id, ref="1.1")
    await _approve(tenant_id, proposal_id, owner_id)

    await ExportService().export(
        tenant_id=tenant_id, workspace_id=workspace_id, export_format=ExportFormat.DOCX
    )

    async with tenant_session(tenant_id) as session:
        stored = await ProposalRepository(session).get_by_id(proposal_id)
    assert stored.status is ProposalStatus.EXPORTED
    # The approver is preserved: the export reviewed nothing.
    assert stored.reviewed_by_id == owner_id
    # And the revision moved, so a reviewer holding the old one is told.
    assert stored.version == 3

    from app.core.exceptions import InvalidTransitionError

    with pytest.raises(InvalidTransitionError):
        await ReviewService().apply_action(
            tenant_id=tenant_id,
            proposal_id=proposal_id,
            actor_id=owner_id,
            actor_role=UserRole.OWNER,
            action=ProposalStatus.REJECTED,
            expected_version=3,
            review_notes="second thoughts",
        )


async def test_a_failed_render_does_not_mark_anything_as_exported(workspace, monkeypatch):
    """Marking first and failing to render would leave answers flagged as
    submitted that never left the building."""
    tenant_id, workspace_id, owner_id = workspace
    proposal_id = await _proposal(tenant_id, workspace_id, ref="1.1")
    await _approve(tenant_id, proposal_id, owner_id)

    import app.services.export_service as module

    def _explode(bundle):
        raise RuntimeError("renderer blew up")

    monkeypatch.setattr(module, "build_response_docx", _explode)

    with pytest.raises(RuntimeError):
        await ExportService().export(
            tenant_id=tenant_id, workspace_id=workspace_id, export_format=ExportFormat.DOCX
        )

    async with tenant_session(tenant_id) as session:
        stored = await ProposalRepository(session).get_by_id(proposal_id)
    assert stored.status is ProposalStatus.APPROVED


async def test_preview_runs_the_gate_without_producing_or_marking_anything(workspace):
    tenant_id, workspace_id, owner_id = workspace
    proposal_id = await _proposal(tenant_id, workspace_id, ref="1.1")
    await _approve(tenant_id, proposal_id, owner_id)

    summary = await ExportService().preview(tenant_id=tenant_id, workspace_id=workspace_id)
    assert summary["answered"] == 1

    async with tenant_session(tenant_id) as session:
        stored = await ProposalRepository(session).get_by_id(proposal_id)
    assert stored.status is ProposalStatus.APPROVED, "preview must not mark anything"


# ---------------------------------------------------------------- isolation
async def test_another_tenants_workspace_is_not_found(workspace, two_tenants):
    tenant_id, workspace_id, owner_id = workspace
    _, other = two_tenants
    proposal_id = await _proposal(tenant_id, workspace_id, ref="1.1")
    await _approve(tenant_id, proposal_id, owner_id)

    from app.core.exceptions import TenantMismatchError

    with pytest.raises(TenantMismatchError):
        await ExportService().export(
            tenant_id=other, workspace_id=workspace_id, export_format=ExportFormat.MATRIX
        )


def _docx_text(content: bytes) -> str:
    """Read the document body straight out of the OOXML package.

    Going through python-docx would test the writer against itself; reading the
    XML is what a different consumer would see.
    """
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        return archive.read("word/document.xml").decode("utf-8")
